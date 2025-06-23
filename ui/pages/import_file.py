from PyQt5.QtWidgets import (
    QWizardPage, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame, QFileDialog, QSizePolicy, QGraphicsDropShadowEffect, QWidget, QMessageBox
)
from PyQt5.QtCore import Qt, QMimeData, pyqtSignal
from PyQt5.QtGui import QIcon, QPixmap, QColor, QCursor
from PyQt5.QtSvg import QSvgWidget
import os
import sys  # for platform checks
from config import SETTINGS_FILE

class DropZone(QFrame):
    fileClicked = pyqtSignal()
    fileDropped = pyqtSignal(str)
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("drop-zone")
        self.setAcceptDrops(True)
        self.setCursor(QCursor(Qt.PointingHandCursor))
        self.setMinimumHeight(120)
        self.setStyleSheet("")
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)
        layout.setSpacing(8)
        layout.setContentsMargins(16, 16, 16, 16)
        # Upload icon
        icon_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../resources/upload.svg'))
        if os.path.exists(icon_path):
            self.upload_icon = QSvgWidget(icon_path)
            self.upload_icon.setFixedSize(32, 32)
            layout.addWidget(self.upload_icon, alignment=Qt.AlignHCenter)
        # Default text
        self.text_label = QLabel("Drag & drop your file here,\nor click 'Browse files…'")
        self.text_label.setStyleSheet("color:#333;font-size:13pt;")
        self.text_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.text_label)
        # Supported formats label
        self.supported_label = QLabel("Supported formats: .xlsx, .csv")
        self.supported_label.setObjectName("supported-label")
        self.supported_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.supported_label)
        # Browse button inside drop zone
        self.browse_button = QPushButton("Browse Files")
        self.browse_button.setObjectName("browse-btn")
        self.browse_button.setCursor(Qt.PointingHandCursor)
        self.browse_button.clicked.connect(self.fileClicked.emit)
        layout.addWidget(self.browse_button, alignment=Qt.AlignCenter)

    def setText(self, text, color="#333"):
        self.text_label.setText(text)
        self.text_label.setStyleSheet(f"color:{color};font-size:13pt;")

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls and self._is_valid_file(urls[0].toLocalFile()):
                self.setProperty("drag", True)
                self.setStyleSheet("border:2px dashed #3897f0;background:#f9f9f9;border-radius:8px;")
                event.acceptProposedAction()
            else:
                self.setStyleSheet("border:2px dashed #d32f2f;background:#f9f9f9;border-radius:8px;")
                event.ignore()
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        self.setProperty("drag", False)
        self.setStyleSheet("")
        event.accept()

    def dropEvent(self, event):
        self.setProperty("drag", False)
        self.setStyleSheet("")
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0]
            file_path = url.toLocalFile()
            if self._is_valid_file(file_path):
                self.fileDropped.emit(file_path)
            else:
                self.fileDropped.emit("")
        event.accept()

    def mousePressEvent(self, event):
        self.fileClicked.emit()

    @staticmethod
    def _is_valid_file(path):
        return path.lower().endswith((".csv", ".xlsx", ".xls"))

class ImportFilePage(QWizardPage):
    def __init__(self, controller):
        super().__init__()
        self.controller = controller
        self.setTitle("")

        # === Restore commented out code ===
        self.setFinalPage(False)
        # self.setButtonLayout([])  # Keep default buttons hidden via wizard options
        self.setCommitPage(True)
        self.setObjectName("import-file-page")
        self.setMinimumSize(0, 0)
        self.setMaximumSize(16777215, 16777215)

        card = QFrame()
        card.setObjectName("card-panel")
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(32, 32, 32, 32)
        card_layout.setSpacing(16)

        title = QLabel("Import NBG or Revolut Statement")
        title.setProperty('role', 'title')
        card_layout.addWidget(title)

        # Drop zone
        self.drop_zone = DropZone()
        self.drop_zone.setObjectName("drop-zone")
        drop_zone_layout = QVBoxLayout(self.drop_zone)
        drop_zone_layout.setContentsMargins(16, 16, 16, 16)
        drop_zone_layout.setSpacing(8)
        # Supported formats label
        self.supported_label = QLabel("Supported formats: .xlsx, .csv")
        self.supported_label.setObjectName("supported-label")
        self.supported_label.setAlignment(Qt.AlignCenter)
        drop_zone_layout.addWidget(self.supported_label)
        # Browse button inside drop zone
        self.browse_button = QPushButton("Browse Files")
        self.browse_button.setObjectName("browse-btn")
        self.browse_button.setCursor(Qt.PointingHandCursor)
        self.browse_button.clicked.connect(self.browse_file) # From click
        drop_zone_layout.addWidget(self.browse_button, alignment=Qt.AlignCenter)
        # Need help? link
        self.help_link = QLabel('<a href="#">Need help?</a>')
        self.help_link.setObjectName("helper-link")
        self.help_link.setAlignment(Qt.AlignCenter)
        self.help_link.setOpenExternalLinks(False)
        self.help_link.linkActivated.connect(self.show_help_modal)
        drop_zone_layout.addWidget(self.help_link)
        card_layout.addWidget(self.drop_zone)

        # File display (filename and clear button)
        self.file_display_widget = QWidget()
        file_display_layout = QHBoxLayout(self.file_display_widget)
        file_display_layout.setContentsMargins(0, 5, 0, 0)
        file_display_layout.setSpacing(6)
        # File type icon
        self.file_icon_label = QLabel()
        self.file_icon_label.setObjectName("file-icon-label")
        self.file_icon_label.setFixedSize(20, 20)
        file_display_layout.addWidget(self.file_icon_label)
        self.file_name_label = QLabel("")
        self.file_name_label.setObjectName("file-name-label")
        self.file_name_label.setAlignment(Qt.AlignCenter)
        self.file_name_label.setCursor(Qt.PointingHandCursor)
        self.file_name_label.mousePressEvent = self.toggle_file_path
        self.showing_full_path = False
        file_display_layout.addWidget(self.file_name_label, 1)
        self.clear_btn = QPushButton("×")
        self.clear_btn.setObjectName("clear-btn")
        self.clear_btn.setFixedSize(20, 20)
        self.clear_btn.setCursor(Qt.PointingHandCursor)
        self.clear_btn.clicked.connect(self.clear_file)
        file_display_layout.addWidget(self.clear_btn)
        self.file_display_widget.hide()
        card_layout.addWidget(self.file_display_widget)

        # Error label
        self.error_label = QLabel("")
        self.error_label.setObjectName("error-label")
        self.error_label.setAlignment(Qt.AlignCenter)
        self.error_label.hide()
        card_layout.addWidget(self.error_label)

        # Spacer
        card_layout.addStretch(1)

        # Navigation Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(12)
        exit_text = "Quit" if sys.platform.startswith('darwin') else "Exit"
        self.exit_button = QPushButton(exit_text)
        self.exit_button.setObjectName("exit-btn")
        self.exit_button.setFixedWidth(100)
        self.exit_button.setFixedHeight(40)
        self.exit_button.setCursor(Qt.PointingHandCursor)
        self.exit_button.clicked.connect(lambda: self.wizard().reject())
        button_layout.addWidget(self.exit_button)
        self.continue_button = QPushButton("Continue")
        self.continue_button.setObjectName("continue-btn")
        self.continue_button.setFixedWidth(100)
        self.continue_button.setFixedHeight(40)
        self.continue_button.setCursor(Qt.PointingHandCursor)
        self.continue_button.clicked.connect(self.validate_and_proceed)
        self.continue_button.setEnabled(False) # Initially disabled
        self.continue_button.setToolTip("Please select a file first")
        button_layout.addStretch(1)
        button_layout.addWidget(self.continue_button)
        card_layout.addLayout(button_layout)

        # State
        self.file_path = None
        self.file_type = None
        self.error_text = ""
        self.last_folder = self.load_last_folder()

        # Connect signals
        # Connect drop event signal
        self.drop_zone.fileDropped.connect(self.handle_file_selected)
        self.drop_zone.fileClicked.connect(self.browse_file) # From click

        # Initial UI state
        self.update_ui_state()
        # === End of restored code ===

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(24, 24, 24, 24)
        main_layout.addWidget(card)
        self.setLayout(main_layout)
        self.setMinimumSize(0, 0)
        self.setMaximumSize(16777215, 16777215)

    def browse_file(self):
        folder = self.last_folder if self.last_folder else os.path.expanduser("~")
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Export File", folder, "CSV/Excel Files (*.csv *.xlsx *.xls)")
        if file_path:
            self.handle_file_selected(file_path)

    def handle_drop_file(self, file_path):
        if file_path:
            self.on_file_selected(file_path)
        else:
            self.show_error("Please select a valid CSV or XLSX file.")

    def on_file_selected(self, file_path):
        if not file_path.lower().endswith((".csv", ".xlsx", ".xls")):
            self.show_error("Please select a valid CSV or XLSX file.")
            return
        self.selected_file_path = file_path
        self.error_text = ""
        self.update_file_display()
        self.validate_file()

    def clear_file(self):
        self.selected_file_path = None
        self.update_file_display()
        self.validate_file()

    def show_error(self, msg):
        self.error_text = msg
        self.error_label.setText(msg)
        self.continue_button.setEnabled(False)
        self.continue_button.setStyleSheet("background:#1976d2;color:#fff;font-weight:600;opacity:0.5;")

    def update_file_display(self):
        if self.selected_file_path:
            self.file_name_label.setText(os.path.basename(self.selected_file_path))
            self.file_display_widget.show()
            self.drop_zone.setText("File selected:", color="#333")
        else:
            self.file_name_label.setText("")
            self.file_display_widget.hide()
            self.drop_zone.setText("Drag & drop your file here,\nor click 'Browse files…'", color="#333")

    def validate_file(self):
        if not self.selected_file_path:
            self.error_label.setText("")
            self.continue_button.setEnabled(False)
            self.continue_button.setStyleSheet("background:#1976d2;color:#fff;font-weight:600;opacity:0.5;")
        elif not self.selected_file_path.lower().endswith((".csv", ".xlsx", ".xls")):
            self.show_error("Please select a valid CSV or XLSX file.")
        else:
            self.error_label.setText("")
            self.continue_button.setEnabled(True)
            self.continue_button.setStyleSheet("background:#1976d2;color:#fff;font-weight:600;opacity:1;")

    def isComplete(self):
        # Allow navigation if a valid file path is set, even if there are warnings
        return self.file_path is not None

    def on_continue(self):
        if not self.isComplete():
            return
        self.wizard().next()

    def go_back(self):
        self.wizard().back()

    def load_last_folder(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, "r") as f:
                    lines = f.readlines()
                for line in lines:
                    if line.startswith("FOLDER:"):
                        return line.strip().split("FOLDER:",1)[1]
            except Exception:
                pass
        return ""

    def save_last_folder(self, folder):
        lines = []
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r") as f:
                for line in f:
                    if line.startswith("TOKEN:"):
                        lines.append(line)
        lines.append(f"FOLDER:{folder}\n")
        with open(SETTINGS_FILE, "w") as f:
            f.writelines(lines)

    def handle_file_selected(self, file_path):
        print(f"[ImportFilePage] handle_file_selected: {file_path}")
        if not file_path:
            return
        self.file_path = file_path
        _, ext = os.path.splitext(file_path)
        # Set file icon
        if ext.lower() == ".csv":
            icon_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../resources/csv_icon.png'))
        else:
            icon_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../resources/excel_icon.png'))
        if os.path.exists(icon_path):
            self.file_icon_label.setPixmap(QPixmap(icon_path).scaled(20, 20, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            self.file_icon_label.clear()
        # Validation: check if file can be opened and has at least 1 data row
        try:
            if ext.lower() == ".csv":
                import csv
                with open(file_path, newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader, None)
                    first_row = next(reader, None)
                    if not headers:
                        raise ValueError("File appears empty or missing headers.")
                    if not first_row:
                        self.error_text = f"Found columns: {', '.join(headers)}. No data rows found."
                    else:
                        self.error_text = f"Found columns: {', '.join(headers)}. Preview OK."
            elif ext.lower() in [".xlsx", ".xls"]:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet = wb.active
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                first_row = next(sheet.iter_rows(min_row=2, max_row=2), None)
                if not headers:
                    raise ValueError("Excel file appears empty or missing headers.")
                if not first_row:
                    self.error_text = f"Found columns: {', '.join(str(h) for h in headers if h)}. No data rows found."
                else:
                    self.error_text = f"Found columns: {', '.join(str(h) for h in headers if h)}. Preview OK."
            else:
                self.error_text = "Invalid file type. Please select a CSV or Excel file."
                self.file_path = None
        except Exception as e:
            self.error_text = f"Error reading file: {str(e)}"
            print(f"[ImportFilePage] Validation error: {self.error_text}")
            self.file_path = None
        else:
            self.file_type = 'csv' if ext.lower() == '.csv' else 'excel'
        self.update_ui_state()
        self.completeChanged.emit() # Notify wizard about completeness change

    def update_ui_state(self):
        if self.file_path:
            self.drop_zone.hide()
            self.drop_zone.browse_button.hide()
            self.file_name_label.setText(os.path.basename(self.file_path))
            self.file_display_widget.show()
            self.continue_button.setEnabled(True)
            self.error_label.hide()
        else:
            self.drop_zone.show()
            self.drop_zone.browse_button.show()
            self.file_display_widget.hide()
            self.continue_button.setEnabled(False)
            if self.error_text:
                self.error_label.setText(self.error_text)
                self.error_label.show()
            else:
                self.error_label.hide()

    def validate_and_proceed(self):
        if self.isComplete():
            # Store file path in controller or shared state if needed later
            # For now, just proceed
            print(f"[ImportFilePage] Proceeding with file: {self.file_path}")
            # Optionally: Trigger file processing/validation in controller here
            # self.controller.set_import_file(self.file_path, self.file_type)
            self.wizard().next()
        else:
            print("[ImportFilePage] Validation failed. Cannot proceed.")
            if not self.error_text:
                self.error_text = "Please select a valid file."
                self.update_ui_state()

    def initializePage(self):
        print(f"[Wizard] initializePage called for page id {self.wizard().currentId()} ({type(self).__name__})")
        # Reset state when page is shown
        # self.clear_file() # Optional: uncomment to always clear file on revisit
        self.update_ui_state()

    def cleanupPage(self):
        print(f"[Wizard] cleanupPage called for page id {self.wizard().currentId()} ({type(self).__name__})")
        # Called when leaving the page
        pass

    def nextId(self):
        # Determine the next page ID based on logic if needed
        # For now, linear progression
        print("[Wizard] nextId called. Next page id: 1")
        return 1 # Assuming YNABAuthPage is always next

    def show_help_modal(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("How to format your NBG/Revolut export")
        msg.setText(
            """
To import your file, export your transactions from NBG or Revolut as .xlsx or .csv.\n\n
- The file should have columns like Date, Description, Amount, etc.\n
- Example (CSV):\n
Date,Description,Amount\n
2025-04-01,Supermarket,-20.00\n
2025-04-02,Coffee,-3.50\n\n
For more details, see the documentation or contact support.
"""
        )
        msg.setIcon(QMessageBox.Information)
        msg.exec_()

    def toggle_file_path(self, event):
        if not self.file_path:
            return
        if self.showing_full_path:
            self.file_name_label.setText(os.path.basename(self.file_path))
            self.showing_full_path = False
        else:
            self.file_name_label.setText(self.file_path)
            self.showing_full_path = True
